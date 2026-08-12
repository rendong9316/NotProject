"""
CSI 300 Weight Validation Tool
Calculates component weights using: free_float_market_cap = price * float_share
Compares with official CSIndex weights from csindex.com
"""
from __future__ import annotations

import time
import sqlite3
import pandas as pd
import baostock as bs
import akshare as ak
import requests
from io import BytesIO
import xlrd
from pathlib import Path
import openpyxl
ROOT = Path(__file__).resolve().parents[2]


SNAPSHOT_DATE = "2026-07-31"
OUTPUT = ROOT / "evidence" / "csi300_weight_validation_20260731.xlsx"
DB_PATH = ROOT / "data" / "csi300_2010_present.sqlite"


def to_bs(code: str) -> str:
    code = str(code).zfill(6)
    return f"sh.{code}" if code.startswith(("5", "6", "68")) else f"sz.{code}"


def fix_name(raw_name: str) -> str:
    """SQLite stores GBK bytes as UTF-8 strings; fix decoding."""
    try:
        return raw_name.encode("utf-8").decode("gbk")
    except Exception:
        return raw_name


def main():
    t0 = time.time()

    # ── 1. Official weights + names from csindex ────────────────
    print("Fetching official weights from csindex.com...")
    url = ("https://oss-ch.csindex.com.cn/static/html/csindex/"
           "public/uploads/file/autofile/closeweight/000300closeweight.xls")
    r = requests.get(url, timeout=30)
    book = xlrd.open_workbook(file_contents=r.content)
    sheet = book.sheet_by_index(0)
    # xlrd already decodes GBK correctly for us
    official_map = {}   # code -> weight
    official_names = {} # code -> Chinese name
    for i in range(1, sheet.nrows):
        code = str(int(sheet.cell_value(i, 4))).zfill(6)
        official_map[code] = float(sheet.cell_value(i, 9))
        official_names[code] = str(sheet.cell_value(i, 5))
    print(f"  {len(official_map)} stocks, date={sheet.cell_value(1, 0)}")

    # ── 2. Price data from SQLite ───────────────────────────────
    print("Loading price data...")
    conn = sqlite3.connect(DB_PATH)
    df_price = pd.read_sql(
        f'SELECT stock_code, close_raw FROM daily_universe '
        f'WHERE date = "{SNAPSHOT_DATE}" AND is_tradeable = 1',
        conn,
    )
    df_security = pd.read_sql("SELECT stock_code, name FROM security", conn)
    conn.close()
    # Fix GBK->UTF-8 encoding issue in SQLite
    name_map = {
        str(row["stock_code"]).zfill(6): fix_name(row["name"])
        for _, row in df_security.iterrows()
    }
    print(f"  {len(df_price)} stocks with prices")

    # ── 3. Share data from Baostock ─────────────────────────────
    print("Fetching share data from Baostock...")
    bs.login()
    share_data = {}
    for i, (_, row) in enumerate(df_price.iterrows()):
        code = str(row["stock_code"]).zfill(6)
        bs_code = to_bs(code)
        found = False
        for year in [2026, 2025, 2024, 2023]:
            for quarter in [4, 3, 2, 1]:
                rs = bs.query_profit_data(code=bs_code, year=year, quarter=quarter)
                if rs.error_code == "0":
                    df_q = rs.get_data()
                    if not df_q.empty:
                        ts, ls = df_q["totalShare"].values[0], df_q["liqaShare"].values[0]
                        if pd.notna(ts) and pd.notna(ls):
                            try:
                                ts_f, ls_f = float(ts), float(ls)
                                if ts_f > 0:
                                    share_data[code] = {
                                        "total": ts_f, "float": ls_f,
                                        "ratio": ls_f / ts_f * 100,
                                        "stat": str(df_q["statDate"].values[0]),
                                    }
                                    found = True
                                    break
                            except Exception:
                                pass
            if found:
                break
        if (i + 1) % 50 == 0:
            print(f"  Progress: {i+1}/{len(df_price)}")
    bs.logout()
    print(f"  {len(share_data)}/{len(df_price)} stocks with share data")

    # ── 4. Calculate weights ────────────────────────────────────
    print("Calculating weights...")
    results = []
    for _, row in df_price.iterrows():
        code = str(row["stock_code"]).zfill(6)
        if code not in share_data:
            continue
        sd = share_data[code]
        results.append({
            "date": SNAPSHOT_DATE,
            "code": code,
            "name": name_map.get(code, official_names.get(code, "")),
            "close": row["close_raw"],
            "total": sd["total"],
            "float": sd["float"],
            "float_ratio": sd["ratio"],
            "stat": sd["stat"],
            "fmv": sd["float"] * row["close_raw"],
        })

    rdf = pd.DataFrame(results)
    total_fmv = rdf["fmv"].sum()
    rdf["calc_weight"] = (rdf["fmv"] / total_fmv * 100).round(4)
    rdf["official_weight"] = rdf["code"].map(official_map)
    rdf["diff_pp"] = (rdf["calc_weight"] - rdf["official_weight"]).round(4)
    rdf["abs_diff_pp"] = rdf["diff_pp"].abs()
    rdf["official_rank"] = rdf["official_weight"].rank(ascending=False, method="min").astype(int)
    rdf["calc_rank"] = rdf["calc_weight"].rank(ascending=False, method="min").astype(int)
    rdf["rank_diff"] = rdf["calc_rank"] - rdf["official_rank"]

    print(f"  Calculated for {len(rdf)} stocks")
    print(f"  Mean abs diff: {rdf['abs_diff_pp'].mean():.4f}pp")
    print(f"  Median abs diff: {rdf['abs_diff_pp'].median():.4f}pp")
    print(f"  Max abs diff: {rdf['abs_diff_pp'].max():.4f}pp")
    print(f"  Correlation: {rdf['calc_weight'].corr(rdf['official_weight']):.4f}")

    # ── 5. Write Excel ──────────────────────────────────────────
    print(f"\nWriting Excel: {OUTPUT}")
    wb = openpyxl.Workbook()

    # Sheet 1: Calculated Weights
    ws1 = wb.active
    ws1.title = "Calculated Weights"
    ws1.append(["Date", "Constituent Code", "Constituent Name", "Weight(%)"])
    for _, r in rdf.sort_values("calc_weight", ascending=False).iterrows():
        ws1.append([r["date"], r["code"], r["name"], round(float(r["calc_weight"]), 4)])

    # Sheet 2: Comparison
    ws2 = wb.create_sheet("Comparison")
    ws2.append(["Constituent Code", "Constituent Name",
                "Calculated Weight(%)", "Official Weight(%)",
                "Diff(pp)", "Abs Diff(pp)",
                "Official Rank", "Calculated Rank", "Rank Diff",
                "Float Ratio(%)", "Stat Date"])
    for _, r in rdf.sort_values("official_weight", ascending=False).iterrows():
        ws2.append([
            r["code"], r["name"],
            round(float(r["calc_weight"]), 4), round(float(r["official_weight"]), 4),
            round(float(r["diff_pp"]), 4), round(float(r["abs_diff_pp"]), 4),
            int(r["official_rank"]), int(r["calc_rank"]), int(r["rank_diff"]),
            round(float(r["float_ratio"]), 2), r["stat"],
        ])

    # Sheet 3: Full Details
    ws3 = wb.create_sheet("Full Details")
    headers = ["date", "code", "name", "close", "total", "float",
               "float_ratio", "stat", "fmv",
               "calc_weight", "official_weight", "diff_pp", "abs_diff_pp",
               "official_rank", "calc_rank", "rank_diff"]
    ws3.append(headers)
    for _, r in rdf.iterrows():
        ws3.append([r[h] for h in headers])

    wb.save(OUTPUT)
    print(f"Done in {time.time()-t0:.1f}s")

    # ── 6. Verify ───────────────────────────────────────────────
    print(f"\nVerification ({OUTPUT}):")
    wb2 = openpyxl.load_workbook(OUTPUT)
    ws_v = wb2["Calculated Weights"]
    for row in range(2, 7):
        print(f"  {ws_v.cell(row, 2).value}: {ws_v.cell(row, 3).value} weight={ws_v.cell(row, 4).value}")


if __name__ == "__main__":
    main()