from pathlib import Path
"""按官网公告校准 xlsx 生效日期列：11 个定期批次(周五收市后生效->下一交易日) + 武钢批次(官网明确 2-14 退市之日起)"""
import sys, datetime, hashlib, shutil
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
ROOT = Path(__file__).resolve().parents[2]


SRC = ROOT / "CSI300_remake_report.xlsx"
BAK = ROOT / "evidence" / "CSI300_remake_report_E校准时备份.xlsx"

SHIFT = {
    '2021-06-11': '2021-06-15', '2021-12-10': '2021-12-13',
    '2022-06-10': '2022-06-13', '2022-12-09': '2022-12-12',
    '2023-06-09': '2023-06-12', '2023-12-08': '2023-12-11',
    '2024-06-14': '2024-06-17', '2024-12-13': '2024-12-16',
    '2025-06-13': '2025-06-16', '2025-12-12': '2025-12-15',
    '2026-06-12': '2026-06-15',
}
WGANG = '2017-02-20'  # -> '2017-02-14'
NEW_REMARK = '武钢股份退市(宝武合并),官网公告4794:自2017-02-14退市之日起调整'

def sha(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1 << 20), b''):
            h.update(chunk)
    return h.hexdigest()[:16]

print('before sha256:', sha(SRC))
shutil.copy2(SRC, BAK)
print('backup ->', BAK)

wb = load_workbook(SRC)
ws = wb.active
n = {'shift': 0, 'wgang': 0}
for r in range(2, ws.max_row + 1):
    e = ws.cell(r, 5).value
    if not isinstance(e, datetime.datetime):
        continue
    key = e.strftime('%Y-%m-%d')
    if key in SHIFT:
        nd = datetime.datetime.strptime(SHIFT[key], '%Y-%m-%d')
        ws.cell(r, 5).value = nd
        n['shift'] += 1
    elif key == WGANG:
        ws.cell(r, 5).value = datetime.datetime(2017, 2, 14)
        ws.cell(r, 6).value = NEW_REMARK
        n['wgang'] += 1
wb.save(SRC)
print('rows changed: shift(定期平移)', n['shift'], '| 武钢', n['wgang'])
print('after sha256:', sha(SRC))