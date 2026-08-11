"""把 2010-2019 沪深300定期调整名单追加到 CSI300_remake_report.xlsx"""
import sys, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from datetime import datetime

# 生效日期映射（公告日 -> 生效日）
EFF_DATE = {
    '2009-12-14': '2010-01-04',
    '2010-06-17': '2010-07-01',
    '2010-12-13': '2011-01-04',
    '2011-06-13': '2011-07-01',
    '2011-12-12': '2012-01-04',
    '2012-06-11': '2012-07-02',
    '2012-12-17': '2013-01-04',
    '2013-06-17': '2013-07-01',
    '2013-12-02': '2013-12-16',
    '2014-06-03': '2014-06-16',
    '2014-12-01': '2014-12-15',
    '2015-06-01': '2015-06-15',
    '2015-11-30': '2015-12-14',
    '2016-05-30': '2016-06-13',
    '2016-11-28': '2016-12-12',
    '2017-05-31': '2017-06-12',
    '2017-11-27': '2017-12-11',
    '2018-05-28': '2018-06-11',
    '2018-12-03': '2018-12-17',
    '2019-06-03': '2019-06-17',
    '2019-12-02': '2019-12-16',
}

data = json.load(open('scratch/hs300_adjustments_full.json', encoding='utf-8'))

# 已有 2019-12-16，跳过
rows = []
for pub_date, rec in data.items():
    eff = EFF_DATE[pub_date]
    if eff == '2019-12-16':
        continue
    for r, a in zip(rec['remove'], rec['add']):
        rows.append((r['code'], r['name'], a['code'], a['name'], eff))

# 降序
rows.sort(key=lambda x: x[4], reverse=True)
print('新增行数:', len(rows))
for eff in sorted(set(r[4] for r in rows), reverse=True):
    n = sum(1 for r in rows if r[4] == eff)
    print(eff, n)

wb = openpyxl.load_workbook('CSI300_remake_report.xlsx')
ws = wb['Sheet1']
next_row = ws.max_row + 1
for code_out, name_out, code_in, name_in, eff in rows:
    ws.cell(row=next_row, column=1, value=code_out)
    ws.cell(row=next_row, column=2, value=name_out)
    ws.cell(row=next_row, column=3, value=code_in)
    ws.cell(row=next_row, column=4, value=name_in)
    c = ws.cell(row=next_row, column=5, value=datetime.strptime(eff, '%Y-%m-%d'))
    c.number_format = 'YYYY-MM-DD'
    next_row += 1

wb.save('CSI300_remake_report.xlsx')
print('已保存 CSI300_remake_report.xlsx')
