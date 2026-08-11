"""批量补入临时调整到 CSI300_remake_report.xlsx"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from datetime import datetime

# (生效日, 调出代码, 调出名称, 调入代码, 调入名称, 备注)
TEMP_ADJ = [
    ('2010-07-29', '600685', '广船国际', '601288', '农业银行', '大市值IPO快速纳入'),
    ('2011-08-23', '600631', '百联股份', '600827', '友谊股份', '吸收合并退市'),
    ('2013-09-18', '000527', '美的电器', '000333', '美的集团', '吸收合并退市'),
    ('2015-01-26', '000562', '宏源证券', '000166', '申万宏源', '吸收合并退市(官网附件缺失,调入据公开信息推断)'),
    ('2015-05-20', '601299', '中国北车', None, None, '吸收合并退市(官网附件缺失,调入待核实)'),
    ('2015-05-20', '600832', '东方明珠', None, None, '吸收合并退市(官网附件缺失,调入待核实)'),
    ('2025-03-04', '600837', '海通证券', '601058', '赛轮轮胎', '国泰君安吸收合并海通证券'),
    ('2025-09-05', '601989', '中国重工', '601298', '青岛港', '中国船舶吸收合并中国重工'),
]

wb = openpyxl.load_workbook('CSI300_remake_report.xlsx')
ws = wb['Sheet1']

# 检查是否已有（按生效日+调出代码去重）
existing = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[4] and r[0]:
        existing.add((str(r[4])[:10], str(r[0]).strip()))

added = 0
for eff, co, no, ci, ni, note in TEMP_ADJ:
    if (eff, co) in existing:
        print(f'跳过(已存在): {eff} {co} {no}')
        continue
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=co)
    ws.cell(row=next_row, column=2, value=no)
    if ci:
        ws.cell(row=next_row, column=3, value=ci)
        ws.cell(row=next_row, column=4, value=ni)
    else:
        ws.cell(row=next_row, column=3, value='(待核实)')
        ws.cell(row=next_row, column=4, value='(官网附件缺失)')
    c = ws.cell(row=next_row, column=5, value=datetime.strptime(eff, '%Y-%m-%d'))
    c.number_format = 'YYYY-MM-DD'
    ws.cell(row=next_row, column=6, value=note)
    existing.add((eff, co))
    added += 1
    print(f'已补: {eff} 调出{co} {no} -> 调入{ci} {ni}  [{note}]')

wb.save('CSI300_remake_report.xlsx')
print(f'\n共新增 {added} 条，总行数 {ws.max_row}')
